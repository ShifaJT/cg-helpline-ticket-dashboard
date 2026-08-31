import streamlit as st
import pandas as pd
import numpy as np
import io
import html
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CG HELPLINE — TICKET INFLOW & RESOLUTION DASHBOARD
# Streamlit version
# ============================================================

st.set_page_config(
    page_title="CG Helpline Ticket Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================
# CSS - KEEP THE DASHBOARD LOOK CLOSE TO THE EXCEL VERSION
# ============================================================

st.markdown("""
<style>
.stApp {
    background-color: #f5f7fb;
}

.block-container {
    padding-top: 1rem;
    padding-left: 2rem;
    padding-right: 2rem;
    max-width: 1600px;
}

.dashboard-header {
    background: #17324d;
    color: white;
    padding: 18px 24px;
    border-radius: 8px;
    margin-bottom: 16px;
}

.dashboard-header h1 {
    margin: 0;
    font-size: 25px;
}

.dashboard-header p {
    margin: 5px 0 0 0;
    font-size: 13px;
    opacity: .82;
}

.filter-header {
    background: #2f75b5;
    color: white;
    padding: 9px 14px;
    border-radius: 7px;
    font-weight: 700;
    margin-bottom: 10px;
}

.section-header {
    color: #17324d;
    font-size: 17px;
    font-weight: 800;
    margin-top: 18px;
    margin-bottom: 10px;
    padding-bottom: 7px;
    border-bottom: 2px solid #dfe5eb;
}

.kpi-card {
    color: white;
    padding: 15px 16px;
    border-radius: 9px;
    min-height: 105px;
}

.kpi-label {
    font-size: 13px;
    font-weight: 700;
}

.kpi-value {
    font-size: 28px;
    font-weight: 800;
    margin-top: 7px;
}

.blue {
    background: #2f75b5;
}

.green {
    background: #70ad47;
}

.red {
    background: #c00000;
}

.orange {
    background: #ed7d31;
}

.small-note {
    font-size: 12px;
    color: #667085;
}

.email-box {
    background: white;
    border: 1px solid #d9e1e8;
    border-radius: 8px;
    padding: 8px;
}

div[data-testid="stDataFrame"] {
    border: 1px solid #dfe5eb;
}

</style>
""", unsafe_allow_html=True)


# ============================================================
# REQUIRED RAW COLUMNS
# ============================================================

REQUIRED_COLUMNS = [
    "Ticket ID",
    "Status",
    "Created time",
    "Closed time",
    "Issue Type - L1",
    "Issue Type - L2",
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_headers(df):
    """Clean only whitespace around headers. Do not rename raw columns."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def read_file(uploaded_file):
    """Read CSV/XLSX/XLS."""
    try:
        filename = uploaded_file.name.lower()

        if filename.endswith(".csv"):
            return pd.read_csv(uploaded_file)

        return pd.read_excel(uploaded_file)

    except Exception as e:
        st.error(f"Could not read the file: {e}")
        return None


def read_paste(text):
    """
    Read Excel copy/paste.
    Excel normally pastes tab-separated values.
    """

    if not text or not text.strip():
        return None

    try:
        df = pd.read_csv(
            io.StringIO(text),
            sep="\t",
            dtype=str,
            keep_default_na=False,
        )

        if len(df.columns) == 1:
            df = pd.read_csv(
                io.StringIO(text),
                dtype=str,
                keep_default_na=False,
            )

        return df

    except Exception as e:

        st.error(
            "Could not read the pasted data. "
            "Make sure you copied the complete Excel table including headers."
        )

        return None


def parse_resolution_hours(series):
    """
    Parse the RAW Freshdesk-style 'Resolution time (in hrs)' column.

    The raw export may arrive as:
      - H:MM:SS text
      - pandas Timedelta
      - datetime/time-like values
      - numeric Excel time fractions
      - numeric hours

    The dashboard uses this raw resolution-time field directly.
    It does NOT calculate resolution as RAW Resolution time (in hrs).
    """

    result = pd.Series(np.nan, index=series.index, dtype="float64")

    for idx, value in series.items():

        if pd.isna(value):
            continue

        # Timedelta values
        if isinstance(value, pd.Timedelta):
            result.loc[idx] = value.total_seconds() / 3600
            continue

        # Python timedelta
        if isinstance(value, __import__("datetime").timedelta):
            result.loc[idx] = value.total_seconds() / 3600
            continue

        # Numeric values:
        # If <= 1, treat as Excel fraction of a day.
        # Otherwise treat as hours.
        if isinstance(value, (int, float, np.integer, np.floating)):
            number = float(value)

            if number < 0:
                continue

            if number <= 1:
                result.loc[idx] = number * 24
            else:
                result.loc[idx] = number

            continue

        text = str(value).strip()

        if not text:
            continue

        # H:MM:SS / HH:MM:SS / D days HH:MM:SS
        try:
            td = pd.to_timedelta(text, errors="coerce")

            if not pd.isna(td):
                result.loc[idx] = td.total_seconds() / 3600
                continue
        except Exception:
            pass

        # Numeric text
        try:
            number = float(text)

            if number >= 0:
                if number <= 1:
                    result.loc[idx] = number * 24
                else:
                    result.loc[idx] = number

        except Exception:
            pass

    return result.clip(lower=0)


def parse_duration_hours(series):
    """Parse raw duration values such as H:MM:SS or numeric hours."""
    result = pd.Series(np.nan, index=series.index, dtype="float64")
    import datetime as _dt

    for row_idx, value in series.items():
        if pd.isna(value):
            continue
        if isinstance(value, pd.Timedelta):
            result.loc[row_idx] = value.total_seconds() / 3600
            continue
        if isinstance(value, _dt.timedelta):
            result.loc[row_idx] = value.total_seconds() / 3600
            continue
        if isinstance(value, (int, float, np.integer, np.floating)):
            n = float(value)
            if n >= 0:
                result.loc[row_idx] = n * 24 if n <= 1 else n
            continue
        text = str(value).strip()
        if not text:
            continue
        try:
            td = pd.to_timedelta(text, errors="coerce")
            if not pd.isna(td):
                result.loc[row_idx] = td.total_seconds() / 3600
                continue
        except Exception:
            pass
        try:
            n = float(text)
            if n >= 0:
                result.loc[row_idx] = n * 24 if n <= 1 else n
        except Exception:
            pass

    return result.clip(lower=0)


def get_first_response_hours(data):
    """Find and parse the raw first-response duration column."""
    for col in [
        "First response time (in hrs)",
        "First response time",
        "First response"
    ]:
        if col in data.columns:
            values = parse_duration_hours(data[col])
            return values[values.notna() & (values >= 0)]
    return pd.Series(dtype="float64")


def parse_datetime_series(series):
    """
    Robust datetime parser.
    Handles normal Excel/export timestamps and mixed values.
    """

    result = pd.to_datetime(
        series,
        errors="coerce",
        dayfirst=False,
    )

    # If too many values failed, try dayfirst.
    if result.notna().sum() < max(1, int(len(series) * 0.5)):

        result2 = pd.to_datetime(
            series,
            errors="coerce",
            dayfirst=True,
        )

        if result2.notna().sum() > result.notna().sum():
            result = result2

    return result


def prepare_data(raw):
    """
    Prepare dashboard fields without modifying the raw export columns.
    """

    if raw is None:
        return None

    df = clean_headers(raw)

    # Required column check
    missing = [
        c for c in REQUIRED_COLUMNS
        if c not in df.columns
    ]

    if missing:

        st.error(
            "Required columns are missing:\n\n"
            + ", ".join(missing)
        )

        return None

    # Remove completely blank rows
    df = df.dropna(how="all").copy()

    # Ticket ID is the safest row identifier.
    df["Ticket ID"] = (
        df["Ticket ID"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df = df[df["Ticket ID"] != ""].copy()

    # --------------------------------------------------------
    # STATUS
    # --------------------------------------------------------

    df["Status Clean"] = (
        df["Status"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    # --------------------------------------------------------
    # ISSUE TYPE
    # --------------------------------------------------------

    df["Issue L1"] = (
        df["Issue Type - L1"]
        .fillna("Unspecified")
        .astype(str)
        .str.strip()
    )

    df.loc[
        df["Issue L1"] == "",
        "Issue L1"
    ] = "Unspecified"

    # Issue Type - L2 is kept exactly as supplied in Raw Data,
    # while Issue L2 is a clean internal field used for analysis.
    df["Issue L2"] = (
        df["Issue Type - L2"]
        .fillna("Unspecified")
        .astype(str)
        .str.strip()
    )

    df.loc[
        df["Issue L2"] == "",
        "Issue L2"
    ] = "Unspecified"

    # --------------------------------------------------------
    # CREATED / CLOSED
    # --------------------------------------------------------

    df["Created time"] = parse_datetime_series(
        df["Created time"]
    )

    df["Closed time"] = parse_datetime_series(
        df["Closed time"]
    )

    df["Created Date"] = (
        df["Created time"]
        .dt.normalize()
    )

    # --------------------------------------------------------
    # MONTH
    # --------------------------------------------------------

    df["Month"] = (
        df["Created time"]
        .dt.strftime("%b %Y")
    )

    # --------------------------------------------------------
    # WEEK
    # Monday-Sunday
    # --------------------------------------------------------

    df["Week Start"] = (
        df["Created Date"]
        -
        pd.to_timedelta(
            df["Created Date"].dt.weekday,
            unit="D",
        )
    )

    df["Week End"] = (
        df["Week Start"]
        +
        pd.Timedelta(days=6)
    )

    df["Week"] = (
        df["Week Start"].dt.strftime("%d %b")
        +
        " - "
        +
        df["Week End"].dt.strftime("%d %b %Y")
    )

    # --------------------------------------------------------
    # DAY
    # --------------------------------------------------------

    df["Day"] = (
        df["Created Date"]
        .dt.strftime("%d %b %Y")
    )

    # --------------------------------------------------------
    # CLOSURE / RESOLUTION HOURS
    #
    # IMPORTANT:
    # Use the RAW "Resolution time (in hrs)" column supplied in
    # the uploaded dump. Do NOT derive this from Closed time -
    # Created time, because the raw resolution metric can differ
    # from elapsed calendar time.
    # --------------------------------------------------------

    if "Resolution time (in hrs)" in df.columns:

        df["Closure Hours"] = parse_resolution_hours(
            df["Resolution time (in hrs)"]
        )

        # Only closed tickets are eligible for closure-time metrics.
        df.loc[
            ~df["Status Clean"].eq("CLOSED"),
            "Closure Hours"
        ] = np.nan

    else:

        # If the raw export does not contain the expected field,
        # leave resolution time unavailable rather than silently
        # creating a different metric.
        df["Closure Hours"] = np.nan

    # --------------------------------------------------------
    # SLA BUCKET
    # --------------------------------------------------------

    # Closed-ticket population used throughout the dashboard.
    # Resolution-time validity is handled in Closure Hours above.
    closed_mask = df["Status Clean"].eq("CLOSED")

    df["SLA"] = "Open"

    df.loc[
        closed_mask
        &
        (df["Closure Hours"] <= 24),
        "SLA"
    ] = "Within 24h"

    df.loc[
        closed_mask
        &
        (df["Closure Hours"] > 24),
        "SLA"
    ] = "After 24h"

    # --------------------------------------------------------
    # OPEN AGE
    # --------------------------------------------------------

    now = pd.Timestamp.now()

    open_mask = (
        df["Status Clean"].eq("OPEN")
        &
        df["Created time"].notna()
    )

    df["Open Age Hours"] = np.nan

    df.loc[
        open_mask,
        "Open Age Hours"
    ] = (
        (
            now
            -
            df.loc[open_mask, "Created time"]
        )
        .dt.total_seconds()
        / 3600
    )

    df["Open Age Hours"] = (
        df["Open Age Hours"]
        .clip(lower=0)
    )

    df = add_fcr_fields(df)

    return df



def ensure_analysis_columns(data):
    """
    Backward-compatible safety layer for Streamlit session state.

    If the app code is updated while the browser session still contains
    an older prepared dataframe, rebuild the internal Issue L1/L2 fields
    from the actual raw columns instead of returning an empty table.
    """
    if data is None:
        return data

    data = data.copy()

    # Recreate Issue L1 from the raw field if necessary.
    if "Issue L1" not in data.columns:
        if "Issue Type - L1" in data.columns:
            data["Issue L1"] = (
                data["Issue Type - L1"]
                .fillna("Unspecified")
                .astype(str)
                .str.strip()
            )
        else:
            data["Issue L1"] = "Unspecified"

    data.loc[
        data["Issue L1"].eq(""),
        "Issue L1"
    ] = "Unspecified"

    # Recreate Issue L2 from the raw field if necessary.
    if "Issue L2" not in data.columns:
        if "Issue Type - L2" in data.columns:
            data["Issue L2"] = (
                data["Issue Type - L2"]
                .fillna("Unspecified")
                .astype(str)
                .str.strip()
            )
        else:
            data["Issue L2"] = "Unspecified"

    data.loc[
        data["Issue L2"].eq(""),
        "Issue L2"
    ] = "Unspecified"

    data = add_fcr_fields(data)

    return data




def add_fcr_fields(data):
    """
    FCR proxy based on the information available in the ticket dump.

    The raw dump does not contain a first-contact-resolution flag or a
    complete contact-history field. Therefore:
      FCR = CLOSED + Created date == Closed date.

    This is a SAME-DAY FCR PROXY, not a proven call-history FCR.
    It is intentionally transparent so the dashboard never reports a
    misleading zero simply because Call Type/Inbound Count are blank.
    """
    x = data.copy()

    if "Created time" not in x.columns:
        x["Created time"] = pd.NaT

    if "Closed time" not in x.columns:
        x["Closed time"] = pd.NaT

    closed = x["Status Clean"].eq("CLOSED")

    valid_times = (
        x["Created time"].notna()
        & x["Closed time"].notna()
    )

    same_day = (
        valid_times
        & x["Created time"].dt.normalize().eq(
            x["Closed time"].dt.normalize()
        )
    )

    x["Same Day Resolution"] = closed & same_day

    # Operational FCR proxy:
    # ticket was raised and resolved during the same calendar day.
    x["FCR"] = x["Same Day Resolution"]

    # Keep these fields for visibility/diagnostics.
    if "Call Type" not in x.columns:
        x["Call Type"] = ""

    if "Inbound Count" not in x.columns:
        x["Inbound Count"] = np.nan

    x["Call Type Clean"] = (
        x["Call Type"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    x["Inbound Count Num"] = pd.to_numeric(
        x["Inbound Count"],
        errors="coerce"
    )

    return x


def fcr_stats(data):
    if data is None or data.empty:
        return {
            "fcr": 0,
            "eligible": 0,
            "rate": 0.0,
            "same_day": 0,
            "coverage": 0.0,
        }

    x = add_fcr_fields(data)

    closed = x["Status Clean"].eq("CLOSED")
    valid_closed = (
        closed
        & x["Created time"].notna()
        & x["Closed time"].notna()
    )

    eligible_count = int(valid_closed.sum())
    fcr_count = int(x["FCR"].sum())
    same_day_count = int(x["Same Day Resolution"].sum())

    return {
        "fcr": fcr_count,
        "eligible": eligible_count,
        # FCR % = FCR tickets / ALL tickets raised
        "rate": (
            fcr_count / len(x) * 100
            if len(x) else 0.0
        ),
        "same_day": same_day_count,
        "coverage": (
            valid_closed.sum() / closed.sum() * 100
            if closed.sum() else 0.0
        ),
    }



def fcr_period_summary(data, period_column):
    if data.empty:
        return pd.DataFrame(
            columns=[
                period_column,
                "Tickets Raised",
                "Closed",
                "FCR",
                "FCR %",
                "Same-Day Resolution",
            ]
        )

    rows = []

    for period, group in data.groupby(
        period_column,
        dropna=False,
        sort=False
    ):
        x = fcr_stats(group)
        rows.append({
            period_column: period,
            "Tickets Raised": group["Ticket ID"].nunique(),
            "Closed": int(
                group["Status Clean"].eq("CLOSED").sum()
            ),
            "FCR": x["fcr"],
            "FCR %": round(x["rate"], 1),
            "Same-Day Resolution": x["same_day"],
        })

    result = pd.DataFrame(rows)

    return result


def format_hms(hours):
    """Convert decimal hours to H:MM:SS."""
    if hours is None or pd.isna(hours):
        return "0:00:00"

    total_seconds = max(0, int(round(float(hours) * 3600)))

    h, rem = divmod(total_seconds, 3600)
    m, s = divmod(rem, 60)

    return f"{h}:{m:02d}:{s:02d}"



def excel_percentile(series, percentile):
    """Excel PERCENTILE.INC-style percentile using linear interpolation."""
    values = pd.to_numeric(series, errors="coerce").dropna()
    values = values[values >= 0]

    if values.empty:
        return 0.0

    return float(
        np.percentile(
            values.to_numpy(dtype=float),
            percentile * 100,
            method="linear"
        )
    )


def get_stats(data):

    if data.empty:
        return {
            "raised": 0,
            "closed": 0,
            "open": 0,
            "within": 0,
            "after": 0,
            "rate": 0.0,
            "avg": 0.0,
            "median": 0.0,
            "p90": 0.0,
            "p95": 0.0,
            "p99": 0.0,
            "p90_rank": 0,
            "p95_rank": 0,
            "p99_rank": 0,
            "old_open": 0,
            "first_response_avg": 0.0,
            "first_response_count": 0,
        }

    closed = data[data["Status Clean"].eq("CLOSED")]
    opened = data[data["Status Clean"].eq("OPEN")]

    valid = closed[
        closed["Closure Hours"].notna()
        & (closed["Closure Hours"] >= 0)
    ].copy()

    within = valid[valid["Closure Hours"] <= 24]
    after = valid[valid["Closure Hours"] > 24]

    if not valid.empty:

        avg_hours = float(valid["Closure Hours"].mean())
        median_hours = float(valid["Closure Hours"].median())

        # Match Excel's PERCENTILE.INC methodology.
        p90_hours = excel_percentile(valid["Closure Hours"], 0.90)
        p95_hours = excel_percentile(valid["Closure Hours"], 0.95)
        p99_hours = excel_percentile(valid["Closure Hours"], 0.99)

        n = len(valid)
        p90_rank = int(np.ceil(n * 0.90))
        p95_rank = int(np.ceil(n * 0.95))
        p99_rank = int(np.ceil(n * 0.99))

    else:
        avg_hours = 0.0
        median_hours = 0.0
        p90_hours = 0.0
        p95_hours = 0.0
        p99_hours = 0.0
        p90_rank = 0
        p95_rank = 0
        p99_rank = 0

    old_open = opened[opened["Open Age Hours"] > 72]

    first_response = get_first_response_hours(data)
    first_response_avg = (
        float(first_response.mean())
        if not first_response.empty
        else 0.0
    )

    return {
        "raised": len(data),
        "closed": len(closed),
        "open": len(opened),
        "within": len(within),
        "after": len(after),
        # 24H Closure % = Closed <=24h / ALL tickets raised
        "rate": (
            len(within) / len(data) * 100
            if len(data)
            else 0.0
        ),
        "avg": avg_hours,
        "median": median_hours,
        "p90": p90_hours,
        "p95": p95_hours,
        "p99": p99_hours,
        "p90_rank": p90_rank,
        "p95_rank": p95_rank,
        "p99_rank": p99_rank,
        "old_open": len(old_open),
        "first_response_avg": first_response_avg,
        "first_response_count": len(first_response),
    }


def show_kpi(label, value, colour_class):

    colours = {
        "blue": "#2f75b5",
        "green": "#70ad47",
        "red": "#c00000",
        "orange": "#ed7d31",
    }

    bg = colours.get(
        colour_class,
        "#2f75b5"
    )

    # st.html prevents Streamlit from displaying the HTML itself.
    st.html(
        f"""
        <div style="
            background:{bg};
            color:white;
            padding:15px 16px;
            border-radius:9px;
            min-height:105px;
            width:100%;
        ">

            <div style="
                font-size:13px;
                font-weight:700;
            ">
                {label}
            </div>

            <div style="
                font-size:28px;
                font-weight:800;
                margin-top:7px;
            ">
                {value}
            </div>

        </div>
        """
    )


def summary_by_group(data, group_column):

    if data.empty:
        return pd.DataFrame(
            columns=[
                group_column,
                "Raised",
                "Closed",
                "Open",
                "≤24h",
                ">24h",
                "24h %",
            ]
        )

    rows = []

    for value, x in data.groupby(
        group_column,
        dropna=False,
        sort=False
    ):

        s = get_stats(x)

        rows.append({
            group_column: value,
            "Raised": s["raised"],
            "Closed": s["closed"],
            "Open": s["open"],
            "≤24h": s["within"],
            ">24h": s["after"],
            "24h %": round(
                s["rate"] * 100,
                1
            ),
        })

    return pd.DataFrame(rows)




def issue_analysis(data):
    """
    Issue Type L1 + L2 analysis using unique Ticket ID.
    """
    cols = [
        "Issue Type - L1",
        "Issue Type - L2",
        "Tickets",
        "% of Tickets",
        "Closed",
        "Open",
        "≤24h",
        ">24h",
        "24h %",
    ]

    if data.empty:
        return pd.DataFrame(columns=cols)

    x = ensure_analysis_columns(data)

    x["Issue L1 Analysis"] = (
        x["Issue Type - L1"]
        .fillna("Unspecified")
        .astype(str)
        .str.strip()
    )

    x.loc[
        x["Issue L1 Analysis"].eq(""),
        "Issue L1 Analysis"
    ] = "Unspecified"

    x["Issue L2 Analysis"] = (
        x["Issue Type - L2"]
        .fillna("Unspecified")
        .astype(str)
        .str.strip()
    )

    x.loc[
        x["Issue L2 Analysis"].eq(""),
        "Issue L2 Analysis"
    ] = "Unspecified"

    rows = []

    for (l1, l2), group in x.groupby(
        ["Issue L1 Analysis", "Issue L2 Analysis"],
        dropna=False,
        sort=False
    ):
        s = get_stats(group)

        rows.append({
            "Issue Type - L1": l1,
            "Issue Type - L2": l2,
            "Tickets": group["Ticket ID"].nunique(),
            "% of Tickets": 0,
            "Closed": s["closed"],
            "Open": s["open"],
            "≤24h": s["within"],
            ">24h": s["after"],
            "24h %": round(s["rate"] * 100, 1),
        })

    result = pd.DataFrame(rows)

    total = result["Tickets"].sum()

    if total:
        result["% of Tickets"] = (
            result["Tickets"] / total * 100
        ).round(1)

    return result.sort_values(
        ["Tickets", "Issue Type - L1", "Issue Type - L2"],
        ascending=[False, True, True]
    ).reset_index(drop=True)


def issue_contributors(data):
    """
    Returns separate Top/Low contributor tables for Issue Type L1 and L2.
    """
    empty = pd.DataFrame(
        columns=["Issue Type", "Tickets", "% of Tickets"]
    )

    if data.empty:
        return empty, empty, empty, empty

    data = ensure_analysis_columns(data)

    def contributor_table(df, column):
        if column not in df.columns:
            return (
                pd.DataFrame(columns=["Issue Type", "Tickets", "% of Tickets"]),
                pd.DataFrame(columns=["Issue Type", "Tickets", "% of Tickets"])
            )

        x = (
            df.groupby(column)["Ticket ID"]
            .nunique()
            .reset_index(name="Tickets")
            .rename(columns={column: "Issue Type"})
        )

        x["Issue Type"] = (
            x["Issue Type"]
            .fillna("Unspecified")
            .astype(str)
            .str.strip()
        )

        x.loc[
            x["Issue Type"].eq(""),
            "Issue Type"
        ] = "Unspecified"

        total = x["Tickets"].sum()

        x["% of Tickets"] = (
            (x["Tickets"] / total * 100).round(1)
            if total
            else 0
        )

        top = x.sort_values(
            ["Tickets", "Issue Type"],
            ascending=[False, True]
        ).head(10).reset_index(drop=True)

        low = x.sort_values(
            ["Tickets", "Issue Type"],
            ascending=[True, True]
        ).head(10).reset_index(drop=True)

        return top, low

    l1_top, l1_low = contributor_table(
        data,
        "Issue L1"
    )

    l2_top, l2_low = contributor_table(
        data,
        "Issue L2"
    )

    return l1_top, l1_low, l2_top, l2_low



def group_period_analysis(data, period_column):
    """
    Period-wise Group analysis:
    shows total tickets and the number recorded in CG Helpline,
    Credit Escalation and High Returns/Reattempts.
    """
    if data.empty:
        return pd.DataFrame()

    x = data.copy()

    x["Group Analysis"] = (
        x["Group"]
        .fillna("")
        .astype(str)
        .str.strip()
        .apply(group_bucket)
    )

    rows = []

    for period, g in x.groupby(
        period_column,
        dropna=False,
        sort=False
    ):
        total = g["Ticket ID"].nunique()

        cg = g.loc[
            g["Group Analysis"].eq("CG Helpline"),
            "Ticket ID"
        ].nunique()

        credit = g.loc[
            g["Group Analysis"].eq("Credit Escalation"),
            "Ticket ID"
        ].nunique()

        returns = g.loc[
            g["Group Analysis"].eq("High Returns/Reattempts"),
            "Ticket ID"
        ].nunique()

        other = max(
            total - cg - credit - returns,
            0
        )

        rows.append({
            period_column: period,
            "Tickets Raised": total,
            "CG Helpline": cg,
            "Credit Escalation": credit,
            "High Returns/Reattempts": returns,
            "Other Groups": other,
        })

    return pd.DataFrame(rows)



def group_bucket(value):
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return "Unassigned"
    n = text.lower()
    if "credit" in n and "escalat" in n:
        return "Credit Escalation"
    if "high returns" in n and "reattempt" in n:
        return "High Returns/Reattempts"
    if "cg helpline" in n or n == "cg":
        return "CG Helpline"
    return text

def group_analysis(data):
    if data.empty or "Group" not in data.columns:
        return pd.DataFrame(columns=["Group", "Tickets", "% of Tickets"])
    x = data.copy()
    x["Group Analysis"] = x["Group"].fillna("").astype(str).str.strip().apply(group_bucket)
    result = (x.groupby("Group Analysis")["Ticket ID"].nunique().reset_index(name="Tickets")
              .rename(columns={"Group Analysis":"Group"})
              .sort_values("Tickets", ascending=False).reset_index(drop=True))
    total = result["Tickets"].sum()
    result["% of Tickets"] = ((result["Tickets"] / total * 100).round(1) if total else 0)
    return result

def requested_group_counts(data):
    if data.empty or "Group" not in data.columns:
        return {"cg":0,"credit":0,"returns":0}
    x=data.copy()
    x["Group Analysis"]=x["Group"].fillna("").astype(str).str.strip().apply(group_bucket)
    return {
        "cg": x.loc[x["Group Analysis"].eq("CG Helpline"),"Ticket ID"].nunique(),
        "credit": x.loc[x["Group Analysis"].eq("Credit Escalation"),"Ticket ID"].nunique(),
        "returns": x.loc[x["Group Analysis"].eq("High Returns/Reattempts"),"Ticket ID"].nunique(),
    }


def apply_dashboard_filters(
    data,
    selected_month,
    selected_week,
    selected_day,
    selected_issue
):
    x = data.copy()

    if selected_month != "All":
        x = x[x["Month"].eq(selected_month)]

    if selected_week != "All":
        x = x[x["Week"].eq(selected_week)]

    if selected_day != "All":
        x = x[x["Day"].eq(selected_day)]

    if selected_issue != "All":
        x = x[x["Issue L1"].eq(selected_issue)]

    return x


def split_current_group(data):
    """
    Main dashboard = current CG Helpline queue.
    Non-CG current groups are treated as moved-out tickets per the
    workflow described by the user.
    """
    x = data.copy()

    if "Group" in x.columns:
        x["Group Analysis"] = (
            x["Group"]
            .fillna("")
            .astype(str)
            .str.strip()
            .apply(group_bucket)
        )
    else:
        x["Group Analysis"] = "Unassigned"

    cg = x[x["Group Analysis"].eq("CG Helpline")].copy()
    moved = x[~x["Group Analysis"].eq("CG Helpline")].copy()

    return cg, moved


def moved_group_summary(data):
    if data.empty:
        return pd.DataFrame(
            columns=[
                "Destination Group",
                "Moved Tickets",
                "Closed",
                "Open",
                "Closed ≤24h",
                "Closed >24h",
                "24h %",
            ]
        )

    rows = []

    for group, g in data.groupby(
        "Group Analysis",
        dropna=False
    ):
        s = get_stats(g)

        rows.append({
            "Destination Group": group,
            "Moved Tickets": g["Ticket ID"].nunique(),
            "Closed": s["closed"],
            "Open": s["open"],
            "Closed ≤24h": s["within"],
            "Closed >24h": s["after"],
            "24h %": round(s["rate"] * 100, 1),
        })

    return (
        pd.DataFrame(rows)
        .sort_values("Moved Tickets", ascending=False)
        .reset_index(drop=True)
    )


def moved_period_summary(data, period_column):
    if data.empty:
        return pd.DataFrame(
            columns=[
                period_column,
                "Moved Tickets",
                "Closed",
                "Open",
                "Closed ≤24h",
                "Closed >24h",
                "24h %",
            ]
        )

    rows = []

    for period, g in data.groupby(
        period_column,
        dropna=False,
        sort=False
    ):
        s = get_stats(g)

        rows.append({
            period_column: period,
            "Moved Tickets": g["Ticket ID"].nunique(),
            "Closed": s["closed"],
            "Open": s["open"],
            "Closed ≤24h": s["within"],
            "Closed >24h": s["after"],
            "24h %": round(s["rate"] * 100, 1),
        })

    return pd.DataFrame(rows)



def filter_description(
    month,
    week,
    day,
    issue
):

    parts = []

    if month != "All":
        parts.append(f"Month: {month}")

    if week != "All":
        parts.append(f"Week: {week}")

    if day != "All":
        parts.append(f"Day: {day}")

    if issue != "All":
        parts.append(f"Issue Type: {issue}")

    if not parts:
        return "All tickets"

    return " | ".join(parts)



def html_escape(value):
    return html.escape(str(value), quote=True)


def html_table(headers, rows):
    header_html = "".join(f"<th>{html_escape(h)}</th>" for h in headers)
    body_html = ""
    for row in rows:
        body_html += "<tr>" + "".join(
            f"<td>{html_escape(v)}</td>" for v in row
        ) + "</tr>"
    return (
        '<table class="email-table">'
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody>{body_html}</tbody>"
        "</table>"
    )


def make_email_html(data, other_group_data, month, week, day, issue):
    s = get_stats(data)
    scope = (
        "CG Helpline current queue | "
        + filter_description(month, week, day, issue)
    )

    l1_summary = summary_by_group(data, "Issue L1")
    l2_summary = summary_by_group(data, "Issue L2")

    l1_rows = []
    for _, row in l1_summary.head(5).iterrows():
        count = int(row["Raised"])
        share = count / s["raised"] * 100 if s["raised"] else 0
        l1_rows.append([row["Issue L1"], f"{count:,}", f"{share:.1f}%"])
    if not l1_rows:
        l1_rows = [["No L1 issue data available.", "0", "0.0%"]]

    l2_rows = []
    for _, row in l2_summary.head(5).iterrows():
        count = int(row["Raised"])
        share = count / s["raised"] * 100 if s["raised"] else 0
        l2_rows.append([row["Issue L2"], f"{count:,}", f"{share:.1f}%"])
    if not l2_rows:
        l2_rows = [["No L2 issue data available.", "0", "0.0%"]]

    # IMPORTANT:
    # data = CG Helpline only.
    # other_group_data = tickets whose CURRENT raw Group is not CG Helpline.
    # This keeps the dashboard KPIs pure CG Helpline while allowing the
    # email to show the separate transferred/current-other-queue view.
    group_col = (
        "Group Analysis"
        if "Group Analysis" in other_group_data.columns
        else "Group"
    )

    transfer_rows = []

    if group_col in other_group_data.columns:

        transferred = other_group_data.copy()

        if not transferred.empty:

            grouped = transferred.groupby(group_col)

            for grp, grp_df in grouped:

                ticket_count = (
                    grp_df["Ticket ID"].nunique()
                    if "Ticket ID" in grp_df.columns
                    else len(grp_df)
                )

                closed_count = (
                    int(
                        grp_df["Status Clean"]
                        .eq("CLOSED")
                        .sum()
                    )
                    if "Status Clean" in grp_df.columns
                    else 0
                )

                open_count = (
                    int(
                        grp_df["Status Clean"]
                        .eq("OPEN")
                        .sum()
                    )
                    if "Status Clean" in grp_df.columns
                    else max(int(ticket_count) - closed_count, 0)
                )

                within_24 = (
                    int(
                        (
                            grp_df["SLA"]
                            .eq("Closed <=24h")
                        ).sum()
                    )
                    if "SLA" in grp_df.columns
                    else 0
                )

                over_24 = (
                    int(
                        (
                            grp_df["SLA"]
                            .eq("Closed >24h")
                        ).sum()
                    )
                    if "SLA" in grp_df.columns
                    else 0
                )

                transfer_rows.append([
                    grp,
                    f"{int(ticket_count):,}",
                    f"{closed_count:,}",
                    f"{open_count:,}",
                    f"{within_24:,}",
                    f"{over_24:,}"
                ])

            # Highest-volume destination queue first.
            transfer_rows.sort(
                key=lambda row: int(
                    str(row[1]).replace(",", "")
                ),
                reverse=True
            )

    if not transfer_rows:
        transfer_rows = [
            ["No tickets currently in other groups.", "0", "0", "0", "0", "0"]
        ]

    findings_rows = [
        ["Tickets raised", f"{s['raised']:,}", "Total ticket inflow"],
        ["Tickets closed", f"{s['closed']:,}", "Tickets closed"],
        ["Tickets open", f"{s['open']:,}", "Current backlog"],
        ["Closed within 24h", f"{s['within']:,}", "Met 24-hour SLA"],
        ["Closed >24h", f"{s['after']:,}", "Exceeded 24-hour SLA"],
        ["24h closure rate", f"{s['rate']:.1f}%", "Closed ≤24h / Total tickets raised"],
        ["Average resolution", format_hms(s["avg"]), "Average closure time"],
        ["Median resolution", format_hms(s["median"]), "Median closure time"],
        ["Average first response", format_hms(s["first_response_avg"]), "Average first response"],
        ["90% Percentile resolution", format_hms(s["p90"]), "PERCENTILE.INC style"],
        ["95% Percentile resolution", format_hms(s["p95"]), "PERCENTILE.INC style"],
        ["99% Percentile resolution", format_hms(s["p99"]), "PERCENTILE.INC style"],
        ["Open >72h", f"{s['old_open']:,}", "Ageing backlog"],
    ]

    if s["closed"]:
        sla_text = (
            f"Out of {s['closed']:,} closed tickets, {s['within']:,} "
            f"were closed within 24 hours and {s['after']:,} exceeded "
            f"the 24-hour SLA. The 24-hour closure rate was "
            f"{s['rate'] * 100:.1f}%."
        )
    else:
        sla_text = "There are no valid closed tickets for SLA analysis."

    resolution_text = (
        f"Average resolution was {format_hms(s['avg'])}; median was "
        f"{format_hms(s['median'])}; 90% percentile was {format_hms(s['p90'])}; "
        f"95% percentile was {format_hms(s['p95'])}; and 99% percentile was "
        f"{format_hms(s['p99'])}."
    )

    backlog_text = f"There are {s['open']:,} tickets currently open."
    if s["old_open"]:
        backlog_text += (
            f" {s['old_open']:,} are older than 72 hours and should be prioritised."
        )
    else:
        backlog_text += " There are no open tickets older than 72 hours."

    if data["Created time"].notna().any():
        date_text = (
            f"The report covers ticket inflow from "
            f"{data['Created time'].min().strftime('%d %b %Y')} to "
            f"{data['Created time'].max().strftime('%d %b %Y')}."
        )
    else:
        date_text = "The reporting period could not be determined."

    return f"""
<div class="copyable-email" style="background:#fff;color:#1f2937;padding:24px;
border:1px solid #d9e1ea;border-radius:8px;font-family:Arial,Helvetica,sans-serif;
line-height:1.5;max-width:1100px;">
<h2 style="color:#173a5e;margin:0 0 4px 0;">CG Helpline Ticket Performance Update</h2>
<p style="color:#64748b;margin-top:0;">{html_escape(scope)}</p>

<p>Hi Team,</p>
<p>Please find below the CG Helpline ticket performance update for
<b>{html_escape(scope)}</b>.</p>

<h3 style="color:#173a5e;">KEY FINDINGS</h3>
{html_table(["Metric","Value","Comments"], findings_rows)}

<h3 style="color:#173a5e;">TOP CONTRIBUTING ISSUE TYPE — L1</h3>
{html_table(["Issue Type — L1","Tickets","% of Tickets"], l1_rows)}

<h3 style="color:#173a5e;">TOP CONTRIBUTING ISSUE TYPE — L2</h3>
{html_table(["Issue Type — L2","Tickets","% of Tickets"], l2_rows)}

<h3 style="color:#173a5e;">TRANSFERRED / OTHER-GROUP TICKETS</h3>
<p>
Based only on the raw <b>Group</b> field. No movement time is inferred.
The table shows the current destination queue and its closed/open and
24-hour resolution status.
</p>
{html_table(
    [
        "Current Group",
        "Tickets",
        "Closed",
        "Open",
        "Closed ≤24h",
        "Closed >24h"
    ],
    transfer_rows
)}

<h3 style="color:#173a5e;">MANAGEMENT OBSERVATIONS</h3>
<p><b>1. Closure and SLA performance</b><br>{html_escape(sla_text)}</p>
<p><b>2. First response time</b><br>
Average first response time was <b>{html_escape(format_hms(s["first_response_avg"]))}</b>,
based on {s["first_response_count"]:,} tickets with a valid first-response value.</p>
<p><b>3. Resolution-time distribution</b><br>{html_escape(resolution_text)}</p>
<p><b>4. Backlog</b><br>{html_escape(backlog_text)}</p>
<p><b>5. Reporting period</b><br>{html_escape(date_text)}</p>

<h3 style="color:#173a5e;">RECOMMENDED FOCUS</h3>
<ul>
<li>Prioritise ageing open tickets and cases approaching the 72-hour threshold.</li>
<li>Review tickets exceeding the 24-hour SLA and identify recurring issue-type drivers.</li>
<li>Review the 90%, 95% and 99% percentile resolution times for long-tail cases.</li>
<li>Focus on the highest-contributing L1 and L2 issues for process or routing improvements.</li>
</ul>
<p>Regards,<br>CG Helpline</p>
</div>
"""

def make_email_summary(
    data,
    month,
    week,
    day,
    issue
):

    s = get_stats(data)

    scope = (
        "CG Helpline current queue | "
        + filter_description(
            month,
            week,
            day,
            issue
        )
    )

    issue_summary = summary_by_group(
        data,
        "Issue L1"
    )

    top_issue = "No issue-type data available"
    top_issue_count = 0
    top_issue_share = 0.0

    if not issue_summary.empty:
        top = issue_summary.iloc[0]
        top_issue = str(top["Issue L1"])
        top_issue_count = int(top["Raised"])
        top_issue_share = (
            top_issue_count / s["raised"] * 100
            if s["raised"] else 0
        )

    second_issue_text = ""

    if len(issue_summary) > 1:
        second = issue_summary.iloc[1]
        second_issue_text = (
            f"The second-highest contributing issue was "
            f"{second['Issue L1']} with {int(second['Raised']):,} tickets."
        )

    if s["closed"]:
        sla_text = (
            f"Out of {s['closed']:,} closed tickets, {s['within']:,} "
            f"were closed within 24 hours and {s['after']:,} exceeded "
            f"the 24-hour SLA. The resulting 24-hour closure rate was "
            f"{s['rate'] * 100:.1f}%."
        )

        resolution_text = (
            f"The average resolution time was {format_hms(s['avg'])} and "
            f"the median was {format_hms(s['median'])}. The 90% percentile "
            f"was {format_hms(s['p90'])}, meaning 90% of valid closed "
            f"tickets were resolved within this time. The 95% percentile "
            f"was {format_hms(s['p95'])}, meaning 95% of valid closed "
            f"tickets were resolved within this time. The 99% percentile "
            f"was {format_hms(s['p99'])}, meaning 99% were resolved "
            f"within this time. 90% and 99% percentile help highlight the long-tail "
            f"cases that take considerably longer to resolve."
        )
    else:
        sla_text = "There are no valid closed tickets for SLA analysis."
        resolution_text = (
            "Average, median, P90 and 99% Percentile resolution times are unavailable "
            "because there are no valid closed-ticket resolution times."
        )

    backlog_text = (
        f"There are {s['open']:,} tickets currently open."
    )

    if s["old_open"]:
        backlog_text += (
            f" {s['old_open']:,} of these are older than 72 hours "
            f"and should be prioritised."
        )
    else:
        backlog_text += (
            " There are no open tickets older than 72 hours."
        )

    if data["Created time"].notna().any():
        min_date = data["Created time"].min()
        max_date = data["Created time"].max()
        date_text = (
            f"The report covers ticket inflow from "
            f"{min_date.strftime('%d %b %Y')} to "
            f"{max_date.strftime('%d %b %Y')}."
        )
    else:
        date_text = "The reporting period could not be determined."

    # Copy/paste-friendly table for Outlook/Gmail.
    findings = [
        ("Tickets raised", f"{s['raised']:,}", "Total ticket inflow"),
        ("Tickets closed", f"{s['closed']:,}", "Tickets closed"),
        ("Tickets open", f"{s['open']:,}", "Current backlog"),
        ("Closed within 24h", f"{s['within']:,}", "Met 24-hour SLA"),
        ("Closed >24h", f"{s['after']:,}", "Exceeded 24-hour SLA"),
        ("24h closure rate", f"{s['rate']:.1f}%", "SLA adherence"),
        ("Average resolution", format_hms(s["avg"]), "Average closure time"),
        ("Median resolution", format_hms(s["median"]), "Middle resolution time"),
        ("Average first response", format_hms(s["first_response_avg"]), "Average raw first-response time"),
        ("90% Percentile resolution", f"{s['p90']:.2f} hrs", "Actual ticket at the 90% ranked position"),
        ("95% Percentile resolution", f"{s['p95']:.2f} hrs", "Actual ticket at the 95% ranked position"),
        ("99% Percentile resolution", f"{s['p99']:.2f} hrs", "Actual ticket at the 99% ranked position"),
        ("Open >72h", f"{s['old_open']:,}", "Ageing backlog"),
    ]

    w1 = max(len("Metric"), max(len(x[0]) for x in findings)) + 3
    w2 = max(len("Value"), max(len(x[1]) for x in findings)) + 3

    table_lines = [
        f"{'Metric'.ljust(w1)}{'Value'.ljust(w2)}Comments",
        "-" * (w1 + w2 + 65),
    ]

    for metric, value, comment in findings:
        table_lines.append(
            f"{metric.ljust(w1)}{value.ljust(w2)}{comment}"
        )

    findings_table = "\n".join(table_lines)

    top_issue_text = (
        f"{top_issue} was the highest-contributing issue type, with "
        f"{top_issue_count:,} tickets, representing {top_issue_share:.1f}% "
        f"of total inflow."
        if top_issue_count
        else "No issue-type contribution could be identified."
    )

    # Top contributing L1/L2 issues.
    l1_summary = summary_by_group(data, "Issue L1")
    l2_summary = summary_by_group(data, "Issue L2")

    top_l1_lines = [
        f"- {row['Issue L1']}: {int(row['Raised']):,} tickets "
        f"({(int(row['Raised']) / s['raised'] * 100) if s['raised'] else 0:.1f}%)"
        for _, row in l1_summary.head(5).iterrows()
    ] or ["- No L1 issue data available."]

    top_l2_lines = [
        f"- {row['Issue L2']}: {int(row['Raised']):,} tickets "
        f"({(int(row['Raised']) / s['raised'] * 100) if s['raised'] else 0:.1f}%)"
        for _, row in l2_summary.head(5).iterrows()
    ] or ["- No L2 issue data available."]

    # Current tickets outside CG Helpline, based only on raw Group.
    group_col = "Group Analysis" if "Group Analysis" in data.columns else "Group"
    if group_col in data.columns:
        transferred = data[
            data[group_col].fillna("").astype(str).str.strip().ne("")
            & ~data[group_col].fillna("").astype(str).str.strip().str.casefold().eq("cg helpline")
        ]
        transfer_counts = (
            transferred.groupby(group_col)["Ticket ID"]
            .nunique()
            .sort_values(ascending=False)
            if not transferred.empty else pd.Series(dtype="int64")
        )
        transfer_lines = [
            f"- {grp}: {int(count):,} tickets"
            for grp, count in transfer_counts.head(10).items()
        ] or ["- No tickets currently in other groups."]
    else:
        transfer_lines = ["- Group column not available in the raw data."]

    email = f"""Subject: CG Helpline Ticket Performance Update – {scope}

Hi Team,

Please find below the CG Helpline ticket performance update for {scope}.

This update provides a concise view of ticket inflow, closure performance, SLA adherence and resolution-time distribution. The key findings are presented in a table for easier review.

KEY FINDINGS

{findings_table}

MANAGEMENT OBSERVATIONS

1. Ticket inflow
{top_issue_text}
{second_issue_text}

2. Closure and SLA performance
{sla_text}

3. First response time
The average first response time was {format_hms(s['first_response_avg'])}, based on {s['first_response_count']:,} tickets with a valid first-response value.

4. Highest-contributing issues

L1 — Top 5:
{chr(10).join(top_l1_lines)}

L2 — Top 5:
{chr(10).join(top_l2_lines)}

5. Transferred / other-group tickets

Tickets currently outside the CG Helpline queue, based only on the raw Group field:
{chr(10).join(transfer_lines)}

6. Resolution-time distribution
{resolution_text}

7. Backlog
{backlog_text}

8. Reporting period
{date_text}

RECOMMENDED FOCUS

• Prioritise ageing open tickets and cases approaching the 72-hour threshold.
• Review tickets exceeding the 24-hour SLA and identify recurring operational or issue-type drivers.
• Review P90, P95 and 99% Percentile resolution times to understand the long-tail cases requiring significantly more time to resolve.
• Focus on the highest-contributing issue types for potential process, routing or knowledge-base improvements.

Overall, the key opportunity is to improve 24-hour SLA adherence while reducing the long-tail resolution time reflected in the 90%, 95% and 99% percentile metrics.

Regards,
CG Helpline
"""

    return email


# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel_report(
    raw_data,
    filtered_data,
    issue_summary,
    month_summary,
    week_summary,
    day_summary,
    email_text,
    filter_text
):
    """
    Creates a professional Excel report containing:
    1. Dashboard
    2. Raw Data - unchanged raw columns/values
    3. Filtered Raw Data
    4. Issue Type Analysis
    5. Monthly Analysis
    6. Weekly Analysis
    7. Daily Analysis
    8. Email Findings

    The raw-data sheet keeps the original raw export columns.
    Calculated helper columns are kept out of Raw Data.
    """

    output = io.BytesIO()

    wb = Workbook()

    # --------------------------------------------------------
    # Theme
    # --------------------------------------------------------

    navy = "17324D"
    blue = "2F75B5"
    green = "70AD47"
    red = "C00000"
    orange = "ED7D31"
    light_blue = "D9EAF7"
    light_green = "E2F0D9"
    light_red = "FCE4D6"
    light_orange = "FCE4D6"
    white = "FFFFFF"
    grey = "F2F4F7"
    dark_grey = "667085"

    thin_grey = Side(
        style="thin",
        color="D9E1E8"
    )

    # --------------------------------------------------------
    # Remove default sheet
    # --------------------------------------------------------

    ws = wb.active
    ws.title = "Dashboard"

    # --------------------------------------------------------
    # Helper functions
    # --------------------------------------------------------

    def style_title(sheet, title, subtitle=None, end_col=8):
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_col
        )

        cell = sheet.cell(1, 1)
        cell.value = title
        cell.fill = PatternFill(
            "solid",
            fgColor=navy
        )
        cell.font = Font(
            color=white,
            bold=True,
            size=18
        )
        cell.alignment = Alignment(
            vertical="center"
        )

        sheet.row_dimensions[1].height = 30

        if subtitle:
            sheet.merge_cells(
                start_row=2,
                start_column=1,
                end_row=2,
                end_column=end_col
            )

            sub = sheet.cell(2, 1)
            sub.value = subtitle
            sub.font = Font(
                color=dark_grey,
                italic=True,
                size=10
            )

    def style_header_row(sheet, row, start_col, end_col):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row, col)
            cell.fill = PatternFill(
                "solid",
                fgColor=navy
            )
            cell.font = Font(
                color=white,
                bold=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = Border(
                bottom=thin_grey
            )

    def add_dataframe(sheet, dataframe, start_row=1, start_col=1):
        """
        Write dataframe with headers and return end row/end col.
        """
        data = dataframe.copy()

        # Convert timestamps to Excel-friendly values
        for col in data.columns:
            if pd.api.types.is_datetime64_any_dtype(data[col]):
                data[col] = data[col].dt.to_pydatetime()

        for j, col in enumerate(data.columns, start_col):
            sheet.cell(
                start_row,
                j,
                str(col)
            )

        style_header_row(
            sheet,
            start_row,
            start_col,
            start_col + len(data.columns) - 1
        )

        for i, row in enumerate(
            data.itertuples(index=False, name=None),
            start_row + 1
        ):
            for j, value in enumerate(
                row,
                start_col
            ):
                if pd.isna(value):
                    value = ""
                sheet.cell(
                    i,
                    j,
                    value
                )

        end_row = start_row + len(data)
        end_col = start_col + len(data.columns) - 1

        return end_row, end_col

    def format_table(sheet, start_row, end_row, start_col, end_col, name):
        if end_row < start_row + 1:
            return

        ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

        table = Table(
            displayName=name,
            ref=ref
        )

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style
        sheet.add_table(table)

    def auto_width(sheet, max_width=32):
        for column_cells in sheet.columns:
            try:
                letter = get_column_letter(
                    column_cells[0].column
                )
            except Exception:
                continue

            max_len = 0

            for cell in column_cells:
                try:
                    value = str(cell.value)
                    max_len = max(
                        max_len,
                        len(value)
                    )
                except Exception:
                    pass

            sheet.column_dimensions[
                letter
            ].width = min(
                max(max_len + 2, 10),
                max_width
            )

    # --------------------------------------------------------
    # DASHBOARD SHEET
    # --------------------------------------------------------

    s = get_stats(filtered_data)

    style_title(
        ws,
        "CG HELPLINE — TICKET INFLOW & RESOLUTION REPORT",
        f"Scope: {filter_text}",
        end_col=8
    )

    # KPI labels
    kpis = [
        ("Tickets Raised", s["raised"], blue),
        ("Closed", s["closed"], green),
        ("Open", s["open"], red),
        ("Closed ≤24h", s["within"], green),
        ("Closed >24h", s["after"], orange),
        ("24h Closure %", f"{s['rate']:.1f}%", green),
        ("Avg Closure Hrs", format_hms(s["avg"]), blue),
        ("Open >72h", s["old_open"], red),
        ("90% Percentile Resolution", format_hms(s["p90"]), blue),
        ("95% Percentile Resolution", format_hms(s["p95"]), orange),
        ("99% Percentile Resolution", format_hms(s["p99"]), red),
    ]

    positions = [
        (4, 1),
        (4, 3),
        (4, 5),
        (4, 7),
        (7, 1),
        (7, 3),
        (7, 5),
        (7, 7),
        (10, 1),
        (10, 4),
        (10, 7),
    ]

    for (label, value, colour), (row, col) in zip(
        kpis,
        positions
    ):
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row,
            end_column=col + 1
        )
        ws.merge_cells(
            start_row=row + 1,
            start_column=col,
            end_row=row + 1,
            end_column=col + 1
        )

        label_cell = ws.cell(row, col)
        label_cell.value = label
        label_cell.fill = PatternFill(
            "solid",
            fgColor=colour
        )
        label_cell.font = Font(
            color=white,
            bold=True
        )
        label_cell.alignment = Alignment(
            horizontal="center"
        )

        value_cell = ws.cell(row + 1, col)
        value_cell.value = value
        value_cell.fill = PatternFill(
            "solid",
            fgColor=colour
        )
        value_cell.font = Font(
            color=white,
            bold=True,
            size=18
        )
        value_cell.alignment = Alignment(
            horizontal="center"
        )

    # Email findings on Dashboard
    ws["A13"] = "IMPORTANT FINDINGS / EMAIL READY"
    ws["A10"].fill = PatternFill(
        "solid",
        fgColor=navy
    )
    ws["A10"].font = Font(
        color=white,
        bold=True,
        size=12
    )

    ws.merge_cells(
        "A14:H26"
    )

    email_cell = ws["A14"]
    email_cell.value = email_text
    email_cell.alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )
    email_cell.fill = PatternFill(
        "solid",
        fgColor="FFFFFF"
    )

    ws["A28"] = "RESOLUTION-TIME PERCENTILE LOGIC"
    ws["A28"].fill = PatternFill(
        "solid",
        fgColor=navy
    )
    ws["A28"].font = Font(
        color=white,
        bold=True,
        size=12
    )

    ws.merge_cells("A29:H31")
    percentile_note = ws["A29"]
    percentile_note.value = (
        "90%, 95% and 99% are calculated from valid CLOSED CG Helpline "
        "ticket resolution times using the Excel PERCENTILE.INC methodology. "
        "The calculation is applied to the filtered closed-ticket population. "
        "Resolution times are displayed as H:MM:SS."
    )
    percentile_note.alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )
    percentile_note.fill = PatternFill(
        "solid",
        fgColor="F2F4F7"
    )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 15

    # --------------------------------------------------------
    # RAW DATA
    # --------------------------------------------------------

    # --------------------------------------------------------
    # RESOLUTION TIME ANALYSIS
    # --------------------------------------------------------
    resolution_ws = wb.create_sheet("Resolution Analysis")

    style_title(
        resolution_ws,
        "CG HELPLINE — RESOLUTION TIME ANALYSIS",
        f"Scope: {filter_text}",
        end_col=5
    )

    resolution_ws["A4"] = "Metric"
    resolution_ws["B4"] = "Value"
    resolution_ws["C4"] = "Definition"

    style_header_row(
        resolution_ws,
        4,
        1,
        3
    )

    resolution_rows = [
        (
            "Closed Tickets",
            s["closed"],
            "Total closed CG Helpline tickets in the selected scope"
        ),
        (
            "Average First Response Time",
            format_hms(s["first_response_avg"]),
            "Average RAW first response time"
        ),
        (
            "Average Closure Time",
            format_hms(s["avg"]),
            "Average of RAW Resolution time (in hrs) for valid closed tickets"
        ),
        (
            "Median Closure Time",
            format_hms(s["median"]),
            "Median resolution time of valid closed tickets"
        ),
        (
            "90% Percentile",
            format_hms(s["p90"]),
            "Excel PERCENTILE.INC(RAW Resolution time (in hrs), 0.90)"
        ),
        (
            "95% Percentile",
            format_hms(s["p95"]),
            "Excel PERCENTILE.INC(RAW Resolution time (in hrs), 0.95)"
        ),
        (
            "99% Percentile",
            format_hms(s["p99"]),
            "Excel PERCENTILE.INC(RAW Resolution time (in hrs), 0.99)"
        ),
        (
            "Closed ≤24h",
            s["within"],
            "Closed tickets resolved within 24 hours"
        ),
        (
            "Closed >24h",
            s["after"],
            "Closed tickets taking more than 24 hours"
        ),
        (
            "24h Closure %",
            f"{s['rate']:.1f}%",
            "Closed ≤24h divided by total closed tickets"
        ),
    ]

    for r, row in enumerate(resolution_rows, start=5):
        for c, value in enumerate(row, start=1):
            resolution_ws.cell(r, c, value)

    style_header_row(
        resolution_ws,
        4,
        1,
        3
    )

    for row in resolution_ws.iter_rows(
        min_row=5,
        max_row=4 + len(resolution_rows),
        min_col=1,
        max_col=3
    ):
        for cell in row:
            cell.border = Border(
                bottom=thin_grey
            )
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    resolution_ws.column_dimensions["A"].width = 28
    resolution_ws.column_dimensions["B"].width = 24
    resolution_ws.column_dimensions["C"].width = 65
    resolution_ws.freeze_panes = "A5"

    raw_ws = wb.create_sheet("Raw Data")

    # Keep original raw columns only.
    raw_export = raw_data.copy()

    # Remove internal calculated columns if present.
    internal_columns = [
        "Status Clean",
        "Issue L1",
        "Issue L2",
        "Created Date",
        "Month",
        "Week Start",
        "Week End",
        "Week",
        "Day",
        "Closure Hours",
        "SLA",
        "Open Age Hours",
    ]

    raw_export = raw_export.drop(
        columns=[
            c for c in internal_columns
            if c in raw_export.columns
        ],
        errors="ignore"
    )

    raw_end_row, raw_end_col = add_dataframe(
        raw_ws,
        raw_export,
        start_row=1
    )

    format_table(
        raw_ws,
        1,
        raw_end_row,
        1,
        raw_end_col,
        "RawDataTable"
    )

    raw_ws.freeze_panes = "A2"
    raw_ws.auto_filter.ref = (
        f"A1:{get_column_letter(raw_end_col)}{raw_end_row}"
    )

    auto_width(
        raw_ws,
        max_width=35
    )

    # --------------------------------------------------------
    # FILTERED RAW DATA
    # --------------------------------------------------------

    filtered_ws = wb.create_sheet(
        "Filtered Raw Data"
    )

    # Keep the original raw columns, but expose the calculated fields needed
    # to audit the dashboard. Raw Data remains untouched.
    filtered_export = filtered_data.drop(
        columns=[
            c for c in internal_columns
            if c in filtered_data.columns
        ],
        errors="ignore"
    ).copy()

    # Explicit ticket-level calculation used by the dashboard:
    # Closed Time - Created Time.
    if "Closure Hours" in filtered_data.columns:
        filtered_export["Resolution Time Used for Dashboard"] = (
            filtered_data["Closure Hours"]
            .apply(format_hms)
        )

    if "SLA" in filtered_data.columns:
        filtered_export["Calculated SLA"] = (
            filtered_data["SLA"]
        )

    for first_col in [
        "First response time (in hrs)",
        "First response time",
        "First response"
    ]:
        if first_col in filtered_data.columns:
            filtered_export["First Response Time Used for Dashboard"] = (
                parse_duration_hours(
                    filtered_data[first_col]
                ).apply(format_hms)
            )
            break

    if "Open Age Hours" in filtered_data.columns:
        filtered_export["Calculated Open Age"] = (
            filtered_data["Open Age Hours"]
            .apply(format_hms)
        )

    end_row, end_col = add_dataframe(
        filtered_ws,
        filtered_export,
        1
    )

    format_table(
        filtered_ws,
        1,
        end_row,
        1,
        end_col,
        "FilteredRawTable"
    )

    filtered_ws.freeze_panes = "A2"

    # Audit note: this is the exact resolution-time calculation used by
    # the dashboard KPIs and percentile calculations.
    note_row = end_row + 3
    filtered_ws.cell(
        note_row,
        1,
        "AUDIT NOTE"
    )
    filtered_ws.cell(
        note_row,
        1
    ).font = Font(
        bold=True,
        color=white
    )
    filtered_ws.cell(
        note_row,
        1
    ).fill = PatternFill(
        "solid",
        fgColor=navy
    )

    filtered_ws.merge_cells(
        start_row=note_row + 1,
        start_column=1,
        end_row=note_row + 3,
        end_column=min(end_col, 8)
    )

    filtered_ws.cell(
        note_row + 1,
        1,
        "Calculated Resolution Time = RAW 'Resolution time (in hrs)' from the uploaded dump. "
        "Only CLOSED tickets with valid Created and Closed timestamps "
        "are included in closure-time, SLA and percentile calculations. "
        "The 90%, 95% and 99% values use Excel PERCENTILE.INC-style "
        "calculation on this calculated resolution-time population. "
        "Displayed ticket-level time is H:MM:SS."
    )
    filtered_ws.cell(
        note_row + 1,
        1
    ).alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )
    filtered_ws.cell(
        note_row + 1,
        1
    ).fill = PatternFill(
        "solid",
        fgColor=grey
    )

    auto_width(
        filtered_ws,
        max_width=35
    )

    # --------------------------------------------------------
    # ISSUE ANALYSIS
    # --------------------------------------------------------

    issue_ws = wb.create_sheet(
        "Issue Analysis"
    )

    style_title(
        issue_ws,
        "ISSUE TYPE-WISE ANALYSIS",
        filter_text,
        end_col=7
    )

    issue_end_row, issue_end_col = add_dataframe(
        issue_ws,
        issue_summary,
        start_row=4
    )

    format_table(
        issue_ws,
        4,
        issue_end_row,
        1,
        issue_end_col,
        "IssueAnalysisTable"
    )

    auto_width(
        issue_ws,
        max_width=32
    )

    # --------------------------------------------------------
    # GROUP ANALYSIS
    # --------------------------------------------------------

    group_ws = wb.create_sheet("Group Analysis")
    style_title(group_ws, "CG / ESCALATION GROUP ANALYSIS", filter_text, end_col=3)
    group_end_row, group_end_col = add_dataframe(group_ws, group_analysis(filtered_data), 4)
    format_table(group_ws, 4, group_end_row, 1, group_end_col, "GroupAnalysisTable")
    auto_width(group_ws, max_width=32)

    # --------------------------------------------------------
    # ISSUE L1 & L2 ANALYSIS
    # --------------------------------------------------------

    issue_detail_ws = wb.create_sheet(
        "Issue L1 & L2 Analysis"
    )

    style_title(
        issue_detail_ws,
        "ISSUE TYPE L1 & L2 ANALYSIS",
        filter_text,
        end_col=9
    )

    issue_detail_end_row, issue_detail_end_col = add_dataframe(
        issue_detail_ws,
        issue_analysis(filtered_data),
        4
    )

    format_table(
        issue_detail_ws,
        4,
        issue_detail_end_row,
        1,
        issue_detail_end_col,
        "IssueL1L2Table"
    )

    auto_width(
        issue_detail_ws,
        max_width=34
    )

    # --------------------------------------------------------
    # TOP / LOW CONTRIBUTORS
    # --------------------------------------------------------

    contrib_ws = wb.create_sheet(
        "Top Low Issues"
    )

    style_title(
        contrib_ws,
        "TOP & LOW CONTRIBUTING ISSUE TYPES",
        filter_text,
        end_col=6
    )

    l1_top_xl, l1_low_xl, l2_top_xl, l2_low_xl = issue_contributors(
        filtered_data
    )

    # L1 top
    contrib_ws["A4"] = "TOP CONTRIBUTING — ISSUE TYPE L1"
    contrib_ws["A4"].font = Font(bold=True, color=white)
    contrib_ws["A4"].fill = PatternFill("solid", fgColor=green)

    l1_top_end_row, l1_top_end_col = add_dataframe(
        contrib_ws,
        l1_top_xl,
        5
    )

    format_table(
        contrib_ws,
        5,
        l1_top_end_row,
        1,
        l1_top_end_col,
        "TopL1IssuesTable"
    )

    # L1 low
    start_l1_low = l1_top_end_row + 3

    contrib_ws.cell(
        start_l1_low,
        1,
        "LOW CONTRIBUTING — ISSUE TYPE L1"
    ).font = Font(bold=True, color=white)

    contrib_ws.cell(
        start_l1_low,
        1
    ).fill = PatternFill("solid", fgColor=orange)

    l1_low_end_row, l1_low_end_col = add_dataframe(
        contrib_ws,
        l1_low_xl,
        start_l1_low + 1
    )

    format_table(
        contrib_ws,
        start_l1_low + 1,
        l1_low_end_row,
        1,
        l1_low_end_col,
        "LowL1IssuesTable"
    )

    # L2 top
    start_l2_top = l1_low_end_row + 3

    contrib_ws.cell(
        start_l2_top,
        1,
        "TOP CONTRIBUTING — ISSUE TYPE L2"
    ).font = Font(bold=True, color=white)

    contrib_ws.cell(
        start_l2_top,
        1
    ).fill = PatternFill("solid", fgColor=green)

    l2_top_end_row, l2_top_end_col = add_dataframe(
        contrib_ws,
        l2_top_xl,
        start_l2_top + 1
    )

    format_table(
        contrib_ws,
        start_l2_top + 1,
        l2_top_end_row,
        1,
        l2_top_end_col,
        "TopL2IssuesTable"
    )

    # L2 low
    start_l2_low = l2_top_end_row + 3

    contrib_ws.cell(
        start_l2_low,
        1,
        "LOW CONTRIBUTING — ISSUE TYPE L2"
    ).font = Font(bold=True, color=white)

    contrib_ws.cell(
        start_l2_low,
        1
    ).fill = PatternFill("solid", fgColor=orange)

    l2_low_end_row, l2_low_end_col = add_dataframe(
        contrib_ws,
        l2_low_xl,
        start_l2_low + 1
    )

    format_table(
        contrib_ws,
        start_l2_low + 1,
        l2_low_end_row,
        1,
        l2_low_end_col,
        "LowL2IssuesTable"
    )

    auto_width(
        contrib_ws,
        max_width=34
    )

    # --------------------------------------------------------
    # MONTHLY GROUP ANALYSIS
    # --------------------------------------------------------

    monthly_group_ws = wb.create_sheet(
        "Monthly Group Analysis"
    )

    style_title(
        monthly_group_ws,
        "MONTH-WISE GROUP BREAKUP",
        filter_text,
        end_col=6
    )

    monthly_group_end_row, monthly_group_end_col = add_dataframe(
        monthly_group_ws,
        group_period_analysis(filtered_data, "Month"),
        4
    )

    format_table(
        monthly_group_ws,
        4,
        monthly_group_end_row,
        1,
        monthly_group_end_col,
        "MonthlyGroupTable"
    )

    auto_width(
        monthly_group_ws,
        max_width=32
    )

    # --------------------------------------------------------
    # WEEKLY GROUP ANALYSIS
    # --------------------------------------------------------

    weekly_group_ws = wb.create_sheet(
        "Weekly Group Analysis"
    )

    style_title(
        weekly_group_ws,
        "WEEK-WISE GROUP BREAKUP",
        filter_text,
        end_col=6
    )

    weekly_group_data = group_period_analysis(
        filtered_data,
        "Week"
    )

    if not weekly_group_data.empty and "Week Start" in filtered_data.columns:

        week_order_xl = (
            filtered_data[["Week", "Week Start"]]
            .drop_duplicates()
            .sort_values("Week Start")
            [["Week"]]
        )

        weekly_group_data = (
            week_order_xl
            .merge(
                weekly_group_data,
                on="Week",
                how="left"
            )
        )

    weekly_group_end_row, weekly_group_end_col = add_dataframe(
        weekly_group_ws,
        weekly_group_data,
        4
    )

    format_table(
        weekly_group_ws,
        4,
        weekly_group_end_row,
        1,
        weekly_group_end_col,
        "WeeklyGroupTable"
    )

    auto_width(
        weekly_group_ws,
        max_width=38
    )

    # --------------------------------------------------------
    # MONTHLY
    # --------------------------------------------------------

    month_ws = wb.create_sheet(
        "Monthly Analysis"
    )

    style_title(
        month_ws,
        "MONTH-WISE ANALYSIS",
        filter_text,
        end_col=7
    )

    month_end_row, month_end_col = add_dataframe(
        month_ws,
        month_summary,
        4
    )

    format_table(
        month_ws,
        4,
        month_end_row,
        1,
        month_end_col,
        "MonthlyAnalysisTable"
    )

    auto_width(
        month_ws,
        max_width=30
    )

    # --------------------------------------------------------
    # WEEKLY
    # --------------------------------------------------------

    week_ws = wb.create_sheet(
        "Weekly Analysis"
    )

    style_title(
        week_ws,
        "WEEK-WISE ANALYSIS",
        filter_text,
        end_col=7
    )

    week_end_row, week_end_col = add_dataframe(
        week_ws,
        week_summary,
        4
    )

    format_table(
        week_ws,
        4,
        week_end_row,
        1,
        week_end_col,
        "WeeklyAnalysisTable"
    )

    auto_width(
        week_ws,
        max_width=35
    )

    # --------------------------------------------------------
    # DAILY
    # --------------------------------------------------------

    day_ws = wb.create_sheet(
        "Daily Analysis"
    )

    style_title(
        day_ws,
        "DAY-WISE ANALYSIS",
        filter_text,
        end_col=8
    )

    day_end_row, day_end_col = add_dataframe(
        day_ws,
        day_summary,
        4
    )

    format_table(
        day_ws,
        4,
        day_end_row,
        1,
        day_end_col,
        "DailyAnalysisTable"
    )

    auto_width(
        day_ws,
        max_width=25
    )

    # --------------------------------------------------------
    # EMAIL FINDINGS
    # --------------------------------------------------------

    email_ws = wb.create_sheet(
        "Email Findings"
    )

    style_title(
        email_ws,
        "IMPORTANT FINDINGS — EMAIL READY",
        filter_text,
        end_col=8
    )

    email_ws.merge_cells(
        "A4:H30"
    )

    email_ws["A4"] = email_text
    email_ws["A4"].alignment = Alignment(
        wrap_text=True,
        vertical="top"
    )
    email_ws["A4"].font = Font(
        size=11
    )

    for col in range(1, 9):
        email_ws.column_dimensions[
            get_column_letter(col)
        ].width = 18

    # --------------------------------------------------------
    # Conditional formatting-like colour coding
    # --------------------------------------------------------

    for sheet in [
        issue_ws,
        month_ws,
        week_ws,
        day_ws
    ]:

        for row in sheet.iter_rows():

            for cell in row:

                if cell.value == "Open":
                    cell.fill = PatternFill(
                        "solid",
                        fgColor="F4CCCC"
                    )

                elif cell.value == "≤24h":
                    cell.fill = PatternFill(
                        "solid",
                        fgColor=light_green
                    )

                elif cell.value == ">24h":
                    cell.fill = PatternFill(
                        "solid",
                        fgColor=light_orange
                    )

    # --------------------------------------------------------
    # Freeze panes
    # --------------------------------------------------------

    for sheet in wb.worksheets:

        if sheet.title not in [
            "Dashboard",
            "Email Findings"
        ]:
            sheet.freeze_panes = "A5" if sheet.title in [
                "Issue Analysis",
                "Monthly Analysis",
                "Weekly Analysis",
                "Daily Analysis"
            ] else "A2"

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    wb.save(output)

    output.seek(0)

    return output.getvalue()


# ============================================================
# HEADER
# ============================================================

st.html(
    """
    <div style="background:linear-gradient(135deg,#17324D 0%,#234F73 60%,#2F75B5 100%);color:white;padding:24px 30px;border-radius:12px;margin-bottom:18px;box-shadow:0 5px 18px rgba(23,50,77,.16);">
        <div style="font-size:12px;font-weight:800;letter-spacing:1.2px;opacity:.78;margin-bottom:6px;">CG HELPLINE • OPERATIONS ANALYTICS</div>
        <div style="font-size:30px;font-weight:800;line-height:1.15;letter-spacing:-.4px;">Ticket Inflow &amp; Resolution Dashboard</div>
        <div style="font-size:14px;margin-top:8px;opacity:.88;">Inflow • SLA performance • backlog ageing • issue mix • escalation flow</div>
    </div>
    """
)


# ============================================================
# SESSION STATE
# ============================================================

if "data" not in st.session_state:

    st.session_state.data = None


# ============================================================
# RAW DATA
# ============================================================

st.markdown(
    '<div class="section-header">RAW DATA</div>',
    unsafe_allow_html=True
)

upload_col, help_col = st.columns(
    [2, 1]
)

with upload_col:

    uploaded_file = st.file_uploader(
        "Upload your raw ticket dump",
        type=[
            "xlsx",
            "xls",
            "csv"
        ],
        help="Use the same raw export you currently use in Excel.",
    )

with help_col:

    st.info(
        "You can upload the raw file OR copy the entire Excel table and paste it below."
    )


paste_data = st.text_area(
    "Paste complete raw dump",
    height=130,
    placeholder=(
        "Copy the complete raw table from Excel, including headers, "
        "and paste it here."
    ),
)


load_button = st.button(
    "Load / Refresh Dashboard",
    type="primary",
)


# ============================================================
# LOAD
# ============================================================

if load_button:

    raw_df = None

    if uploaded_file is not None:

        raw_df = read_file(
            uploaded_file
        )

    elif paste_data.strip():

        raw_df = read_paste(
            paste_data
        )

    if raw_df is not None:

        prepared = prepare_data(
            raw_df
        )

        if prepared is not None:

            st.session_state.data = prepared

            st.success(
                f"Successfully loaded {len(prepared):,} tickets."
            )


# ============================================================
# WAIT FOR DATA
# ============================================================

if st.session_state.data is None:

    st.warning(
        "Please upload or paste your raw ticket dump."
    )

    st.stop()


df = ensure_analysis_columns(
    st.session_state.data
)

# Store the repaired dataframe back into the session so the same
# raw upload works even after the app code has been updated.
st.session_state.data = df

# ============================================================
# CURRENT GROUP SPLIT
# ============================================================

# MAIN DASHBOARD = ONLY CURRENT CG HELPLINE GROUP
# OTHER GROUPS = CURRENTLY WITH DESTINATION QUEUES
if "Group" in df.columns:
    df["_Group Clean"] = (
        df["Group"]
        .fillna("")
        .astype(str)
        .str.strip()
    )
else:
    df["_Group Clean"] = ""

cg_all = df[
    df["_Group Clean"].str.casefold().eq("cg helpline")
].copy()

other_group_all = df[
    ~df["_Group Clean"].str.casefold().eq("cg helpline")
    & df["_Group Clean"].ne("")
].copy()


# ============================================================
# DATA HEALTH CHECK
# ============================================================

invalid_created = df["Created time"].isna().sum()

if invalid_created:

    st.warning(
        f"{invalid_created:,} ticket(s) have an invalid/missing Created time "
        "and will not appear correctly in month/week/day analysis."
    )


# ============================================================
# FILTERS
# ============================================================

st.markdown(
    '<div class="filter-header">FILTERS</div>',
    unsafe_allow_html=True
)

f1, f2, f3, f4 = st.columns(4)


# ------------------------------------------------------------
# MONTH
# ------------------------------------------------------------

# Keep each month only once. The previous version used
# drop_duplicates() on both Month and Created time, which meant
# every different date in the same month created another "Jul 2026"
# entry in the dropdown.
month_values = (
    cg_all[["Month", "Created time"]]
    .dropna(subset=["Month", "Created time"])
    .sort_values("Created time")
    .drop_duplicates(subset=["Month"], keep="first")
    ["Month"]
    .tolist()
)

with f1:

    selected_month = st.selectbox(
        "Month",
        ["All"] + month_values
    )


# ------------------------------------------------------------
# WEEK
# ------------------------------------------------------------

month_filtered = cg_all.copy()

if selected_month != "All":

    month_filtered = month_filtered[
        month_filtered["Month"]
        ==
        selected_month
    ]

week_values = (
    month_filtered[
        ["Week", "Week Start"]
    ]
    .dropna(subset=["Week"])
    .drop_duplicates()
    .sort_values("Week Start")
    ["Week"]
    .tolist()
)

with f2:

    selected_week = st.selectbox(
        "Week",
        ["All"] + week_values
    )


# ------------------------------------------------------------
# DAY
# ------------------------------------------------------------

week_filtered = month_filtered.copy()

if selected_week != "All":

    week_filtered = week_filtered[
        week_filtered["Week"]
        ==
        selected_week
    ]

day_values = (
    week_filtered[
        ["Day", "Created Date"]
    ]
    .dropna(subset=["Day"])
    .drop_duplicates()
    .sort_values("Created Date")
    ["Day"]
    .tolist()
)

with f3:

    selected_day = st.selectbox(
        "Day",
        ["All"] + day_values
    )


# ------------------------------------------------------------
# ISSUE TYPE
# ------------------------------------------------------------

issue_values = sorted(
    cg_all["Issue L1"]
    .dropna()
    .unique()
)

with f4:

    selected_issue = st.selectbox(
        "Issue Type - L1",
        ["All"] + list(issue_values)
    )


# ============================================================
# APPLY FILTERS
# ============================================================

# Main dashboard is CURRENT CG HELPLINE only.
filtered = apply_dashboard_filters(
    cg_all,
    selected_month,
    selected_week,
    selected_day,
    selected_issue
)

# Separate view for tickets currently in other queues.
filtered_other_groups = apply_dashboard_filters(
    other_group_all,
    selected_month,
    selected_week,
    selected_day,
    selected_issue
)


# ============================================================
# FILTER INFO
# ============================================================

st.markdown(
    f"""
    <div class="small-note">
    Showing <b>{filtered["Ticket ID"].nunique():,}</b> current CG Helpline tickets | 
    {filter_description(selected_month, selected_week, selected_day, selected_issue)}
    </div>
    """,
    unsafe_allow_html=True
)


# ============================================================
# SUMMARY / KPI
# ============================================================

s = get_stats(filtered)

st.markdown(
    '<div class="section-header">SUMMARY</div>',
    unsafe_allow_html=True
)

r1 = st.columns(4)

with r1[0]:
    show_kpi(
        "TICKETS RAISED",
        f"{s['raised']:,}",
        "blue"
    )

with r1[1]:
    show_kpi(
        "CLOSED",
        f"{s['closed']:,}",
        "green"
    )

with r1[2]:
    show_kpi(
        "OPEN",
        f"{s['open']:,}",
        "red"
    )

with r1[3]:
    show_kpi(
        "CLOSED ≤24H",
        f"{s['within']:,}",
        "green"
    )

st.write("")

r2 = st.columns(4)

with r2[0]:
    show_kpi(
        "CLOSED >24H",
        f"{s['after']:,}",
        "orange"
    )

with r2[1]:
    show_kpi(
        "24H CLOSURE %",
        f"{s['rate']:.1f}%",
        "green"
    )

with r2[2]:
    show_kpi(
        "AVG CLOSURE HRS",
        format_hms(s["avg"]),
        "blue"
    )

with r2[3]:
    show_kpi(
        "OPEN >72H",
        f"{s['old_open']:,}",
        "red"
    )

st.write("")

# Resolution-time distribution
# 90%/95%/99% percentile are calculated only from valid closed-ticket resolution hours.
r3 = st.columns(3)

with r3[0]:
    show_kpi(
        "90% PERCENTILE RESOLUTION HRS",
        format_hms(s["p90"]),
        "blue"
    )

with r3[1]:
    show_kpi(
        "95% PERCENTILE RESOLUTION HRS",
        format_hms(s["p95"]),
        "orange"
    )

with r3[2]:
    show_kpi(
        "99% PERCENTILE RESOLUTION HRS",
        format_hms(s["p99"]),
        "red"
    )

st.caption(
    "24H Closure % = Closed ≤24h ÷ Total tickets raised. "
    "All percentage KPIs use total tickets raised in the selected population "
    "as the denominator. 90%/95%/99% percentile values are time measures."
)



r4 = st.columns(4)

with r4[0]:
    show_kpi(
        "AVG FIRST RESPONSE HRS",
        format_hms(s["first_response_avg"]),
        "blue"
    )

with r4[1]:
    show_kpi(
        "OVERALL TICKETS",
        f"{s['first_response_count']:,}",
        "green"
    )

st.write("")


# ============================================================
# FCR / FIRST CONTACT RESOLUTION
# ============================================================

st.markdown(
    '<div class="section-header">FCR / FIRST CONTACT RESOLUTION</div>',
    unsafe_allow_html=True
)

fcr = fcr_stats(filtered)

f1, f2, f3, f4 = st.columns(4)

with f1:
    show_kpi(
        "FCR TICKETS",
        f"{fcr['fcr']:,}",
        "blue"
    )

with f2:
    show_kpi(
        "FCR %",
        f"{fcr['rate']:.1f}%",
        "green"
    )

with f3:
    show_kpi(
        "FCR ELIGIBLE",
        f"{fcr['eligible']:,}",
        "orange"
    )

with f4:
    show_kpi(
        "SAME-DAY RESOLUTION",
        f"{fcr['same_day']:,}",
        "green"
    )

st.caption(
    "FCR operational proxy = Closed ticket + populated Call Type + Inbound Count ≤ 1. "
    "FCR % = FCR tickets ÷ Total tickets raised. "
    "The raw export does not contain a dedicated FCR flag/history."
)

if fcr["eligible"] == 0:
    st.warning(
        "No closed tickets have valid Created and Closed timestamps, "
        "so the FCR proxy cannot be calculated."
    )

fcr_week = fcr_period_summary(filtered, "Week")

if not fcr_week.empty:

    st.markdown("**FCR by Week**")

    st.dataframe(
        fcr_week,
        use_container_width=True,
        hide_index=True
    )

# ============================================================
# CG HELPLINE — CURRENT QUEUE
# ============================================================

st.markdown(
    '<div class="section-header">CG HELPLINE — CURRENT QUEUE</div>',
    unsafe_allow_html=True
)

st.caption(
    "The main Summary, SLA, FCR and Issue Type analysis uses ONLY tickets "
    "whose current Group is CG Helpline."
)



# ============================================================
# CURRENTLY IN OTHER GROUPS
# ============================================================


st.markdown("### CURRENT GROUP DISTRIBUTION — RAW GROUP COLUMN")

group_distribution = (
    df.groupby("_Group Clean")["Ticket ID"]
    .nunique()
    .reset_index(name="Tickets")
    .rename(columns={"_Group Clean": "Current Group"})
    .sort_values("Tickets", ascending=False)
    .reset_index(drop=True)
)

st.dataframe(
    group_distribution,
    use_container_width=True,
    hide_index=True
)

st.markdown(
    '<div class="section-header">CURRENT QUEUE MOVEMENT VIEW</div>',
    unsafe_allow_html=True
)

st.caption(
    "This section uses ONLY the raw Group column. "
    "The main dashboard contains only Group = CG Helpline. "
    "Every ticket whose current Group is different from CG Helpline is "
    "shown here as being in an Other Queue. "
    "≤24h / >24h is based on the raw resolution-time field; "
    "24H Closure % = Closed ≤24h ÷ Total tickets in other queues."
)

other = filtered_other_groups.copy()

other_stats = get_stats(other)

other_total = other["Ticket ID"].nunique()

other_within = other_stats["within"]
other_after = other_stats["after"]
other_closed = other_stats["closed"]
other_open = other_stats["open"]

# Other/L3 24H Closure % = Closed <=24h / ALL tickets in other queues
other_rate = (
    other_within / other_total * 100
    if other_total else 0
)

o1, o2, o3, o4 = st.columns(4)

with o1:
    show_kpi(
        "TICKETS IN OTHER QUEUES",
        f"{other_total:,}",
        "orange"
    )

with o2:
    show_kpi(
        "CLOSED",
        f"{other_closed:,}",
        "green"
    )

with o3:
    show_kpi(
        "OPEN",
        f"{other_open:,}",
        "red"
    )

with o4:
    show_kpi(
        "CLOSED ≤24H",
        f"{other_within:,}",
        "blue"
    )

o5, o6, o7 = st.columns(3)

with o5:
    show_kpi(
        "CLOSED >24H",
        f"{other_after:,}",
        "orange"
    )

with o6:
    show_kpi(
        "24H CLOSURE %",
        f"{other_rate:.1f}%",
        "green"
    )

with o7:
    show_kpi(
        "DESTINATION QUEUES",
        f"{other['_Group Clean'].nunique():,}"
        if "_Group Clean" in other.columns
        else "0",
        "blue"
    )

st.caption(
    "Queue Count = number of distinct non-CG Helpline values in the raw Group column. "
    "It is NOT the number of tickets. The ticket count is shown in TICKETS IN OTHER QUEUES."
)

st.markdown("### OTHER QUEUE-WISE BREAKDOWN")

if not other.empty:

    queue_rows = []

    for queue, g in other.groupby(
        "_Group Clean",
        dropna=False
    ):
        s = get_stats(g)

        queue_rows.append({
            "Current Queue": queue,
            "Tickets": g["Ticket ID"].nunique(),
            "Closed": s["closed"],
            "Open": s["open"],
            "Closed ≤24h": s["within"],
            "Closed >24h": s["after"],
            "24h Closure %": round(
                s["rate"] * 100,
                1
            ),
        })

    queue_df = (
        pd.DataFrame(queue_rows)
        .sort_values(
            "Tickets",
            ascending=False
        )
        .reset_index(drop=True)
    )

    q1, q2 = st.columns([1.05, 0.95])

    with q1:
        st.dataframe(
            queue_df,
            use_container_width=True,
            hide_index=True
        )

    with q2:
        import plotly.express as px

        fig_queue = px.bar(
            queue_df.sort_values("Tickets"),
            x="Tickets",
            y="Current Queue",
            orientation="h",
            text="Tickets",
            title="Tickets Currently in Other Queues"
        )

        fig_queue.update_layout(
            height=360,
            margin=dict(
                l=10,
                r=10,
                t=50,
                b=10
            )
        )

        st.plotly_chart(
            fig_queue,
            use_container_width=True
        )

else:
    st.info(
        "No tickets are currently in an Other Queue for the selected filters."
    )

st.markdown("### Other Queues — Month-wise")

other_month = summary_by_group(
    other,
    "Month"
)

if not other_month.empty:
    st.dataframe(
        other_month,
        use_container_width=True,
        hide_index=True
    )

st.markdown("### Other Queues — Week-wise")

other_week = summary_by_group(
    other,
    "Week"
)

if not other_week.empty:
    st.dataframe(
        other_week,
        use_container_width=True,
        hide_index=True
    )

st.markdown("### Other Queues — Day-wise")

other_day = summary_by_group(
    other,
    "Day"
)

if not other_day.empty:
    st.dataframe(
        other_day,
        use_container_width=True,
        hide_index=True
    )


# ============================================================
# ISSUE TYPE-WISE
# ============================================================

st.markdown(
    '<div class="section-header">ISSUE TYPE-WISE BREAKUP — L1 & L2</div>',
    unsafe_allow_html=True
)

issue_detail = issue_analysis(filtered)

if "Issue Type - L2" in filtered.columns:
    l2_source = (
        filtered["Issue Type - L2"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    if l2_source.eq("").all():
        st.warning(
            "Issue Type - L2 is blank in the selected raw data. "
            "The dashboard cannot create an L2 breakdown unless the raw dump contains L2 values."
        )

if not issue_detail.empty:

    st.dataframe(
        issue_detail,
        use_container_width=True,
        hide_index=True
    )

    l1_top, l1_low, l2_top, l2_low = issue_contributors(
        filtered
    )

    st.markdown(
        '<div class="section-header">TOP & LOW CONTRIBUTING ISSUES</div>',
        unsafe_allow_html=True
    )

    # -------------------------
    # L1 VIEW
    # -------------------------

    st.markdown("### Issue Type - L1")

    l1c1, l1c2 = st.columns(2)

    with l1c1:

        st.markdown("**Top contributing L1 issues**")

        st.dataframe(
            l1_top,
            use_container_width=True,
            hide_index=True
        )

        if not l1_top.empty:

            import plotly.express as px

            fig_l1_top = px.bar(
                l1_top.sort_values("Tickets"),
                x="Tickets",
                y="Issue Type",
                orientation="h",
                text="Tickets",
                title="Top L1 Contributors"
            )

            fig_l1_top.update_layout(
                height=380,
                margin=dict(
                    l=10,
                    r=10,
                    t=50,
                    b=10
                )
            )

            st.plotly_chart(
                fig_l1_top,
                use_container_width=True
            )

    with l1c2:

        st.markdown("**Low contributing L1 issues**")

        st.dataframe(
            l1_low,
            use_container_width=True,
            hide_index=True
        )

        if not l1_low.empty:

            import plotly.express as px

            fig_l1_low = px.bar(
                l1_low.sort_values("Tickets"),
                x="Tickets",
                y="Issue Type",
                orientation="h",
                text="Tickets",
                title="Low L1 Contributors"
            )

            fig_l1_low.update_layout(
                height=380,
                margin=dict(
                    l=10,
                    r=10,
                    t=50,
                    b=10
                )
            )

            st.plotly_chart(
                fig_l1_low,
                use_container_width=True
            )

    # -------------------------
    # L2 VIEW
    # -------------------------

    st.markdown("### Issue Type - L2")

    l2c1, l2c2 = st.columns(2)

    with l2c1:

        st.markdown("**Top contributing L2 issues**")

        st.dataframe(
            l2_top,
            use_container_width=True,
            hide_index=True
        )

        if not l2_top.empty:

            import plotly.express as px

            fig_l2_top = px.bar(
                l2_top.sort_values("Tickets"),
                x="Tickets",
                y="Issue Type",
                orientation="h",
                text="Tickets",
                title="Top L2 Contributors"
            )

            fig_l2_top.update_layout(
                height=420,
                margin=dict(
                    l=10,
                    r=10,
                    t=50,
                    b=10
                )
            )

            st.plotly_chart(
                fig_l2_top,
                use_container_width=True
            )

    with l2c2:

        st.markdown("**Low contributing L2 issues**")

        st.dataframe(
            l2_low,
            use_container_width=True,
            hide_index=True
        )

        if not l2_low.empty:

            import plotly.express as px

            fig_l2_low = px.bar(
                l2_low.sort_values("Tickets"),
                x="Tickets",
                y="Issue Type",
                orientation="h",
                text="Tickets",
                title="Low L2 Contributors"
            )

            fig_l2_low.update_layout(
                height=420,
                margin=dict(
                    l=10,
                    r=10,
                    t=50,
                    b=10
                )
            )

            st.plotly_chart(
                fig_l2_low,
                use_container_width=True
            )


# ============================================================
# MONTH-WISE
# ============================================================

st.markdown(
    '<div class="section-header">MONTH-WISE BREAKUP</div>',
    unsafe_allow_html=True
)

month_summary = summary_by_group(
    filtered,
    "Month"
)

if not month_summary.empty:

    month_summary["_sort"] = (
        month_summary["Month"]
        .apply(
            lambda x: pd.to_datetime(
                "01 " + x
            )
        )
    )

    month_summary = (
        month_summary
        .sort_values("_sort")
        .drop(columns="_sort")
    )

    left, right = st.columns(
        [1.1, 0.9]
    )

    with left:

        st.dataframe(
            month_summary,
            use_container_width=True,
            hide_index=True
        )

    with right:

        fig = px.line(
            month_summary,
            x="Month",
            y="Raised",
            markers=True,
            title="Monthly Ticket Inflow"
        )

        fig.update_layout(
            height=350,
            margin=dict(
                l=10,
                r=10,
                t=50,
                b=10
            )
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )



# ============================================================
# MONTH-WISE GROUP MOVEMENT / ALLOCATION
# ============================================================

st.markdown(
    '<div class="section-header">MONTH-WISE GROUP BREAKUP</div>',
    unsafe_allow_html=True
)

month_group = group_period_analysis(
    filtered,
    "Month"
)

if not month_group.empty:

    st.dataframe(
        month_group,
        use_container_width=True,
        hide_index=True
    )


# ============================================================
# WEEK-WISE
# ============================================================

st.markdown(
    '<div class="section-header">WEEK-WISE BREAKUP</div>',
    unsafe_allow_html=True
)

week_summary = summary_by_group(
    filtered,
    "Week"
)

if not week_summary.empty:

    week_order = (
        filtered[
            ["Week", "Week Start"]
        ]
        .drop_duplicates()
        .sort_values("Week Start")
    )

    week_summary = (
        week_order[["Week"]]
        .merge(
            week_summary,
            on="Week",
            how="left"
        )
    )

    left, right = st.columns(
        [1.1, 0.9]
    )

    with left:

        st.dataframe(
            week_summary,
            use_container_width=True,
            hide_index=True
        )

    with right:

        fig = px.line(
            week_summary,
            x="Week",
            y="Raised",
            markers=True,
            title="Weekly Ticket Inflow"
        )

        fig.update_layout(
            height=350,
            margin=dict(
                l=10,
                r=10,
                t=50,
                b=10
            )
        )

        st.plotly_chart(
            fig,
            use_container_width=True
        )



# ============================================================
# WEEK-WISE GROUP MOVEMENT / ALLOCATION
# ============================================================

st.markdown(
    '<div class="section-header">WEEK-WISE GROUP BREAKUP</div>',
    unsafe_allow_html=True
)

week_group = group_period_analysis(
    filtered,
    "Week"
)

if not week_group.empty:

    # Keep the same chronological order as the weekly analysis.
    if "Week Start" in filtered.columns:
        week_order = (
            filtered[["Week", "Week Start"]]
            .drop_duplicates()
            .sort_values("Week Start")
            [["Week"]]
        )

        week_group = (
            week_order
            .merge(
                week_group,
                on="Week",
                how="left"
            )
        )

    st.dataframe(
        week_group,
        use_container_width=True,
        hide_index=True
    )


# ============================================================
# DAY-WISE
# ============================================================

st.markdown(
    '<div class="section-header">DAY-WISE BREAKUP</div>',
    unsafe_allow_html=True
)

day_summary = summary_by_group(
    filtered,
    "Day"
)

if not day_summary.empty:

    day_summary["_sort"] = (
        pd.to_datetime(
            day_summary["Day"]
        )
    )

    day_summary = (
        day_summary
        .sort_values("_sort")
        .drop(columns="_sort")
    )

    day_summary.insert(
        1,
        "Day Name",
        pd.to_datetime(
            day_summary["Day"]
        ).dt.strftime("%A")
    )

    st.dataframe(
        day_summary,
        use_container_width=True,
        hide_index=True
    )


st.markdown(
    """
    <style>
    .copyable-email .email-table {
        border-collapse: collapse;
        width: 100%;
        margin: 10px 0 20px 0;
        font-family: Arial, Helvetica, sans-serif;
        font-size: 13px;
    }
    .copyable-email .email-table th {
        background: #173a5e;
        color: white;
        border: 1px solid #cbd5e1;
        padding: 8px 10px;
        text-align: left;
    }
    .copyable-email .email-table td {
        border: 1px solid #cbd5e1;
        padding: 8px 10px;
        background: white;
        color: #1f2937;
        vertical-align: top;
    }
    .copyable-email .email-table tr:nth-child(even) td {
        background: #f8fafc;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# ============================================================
# EMAIL FINDINGS
# ============================================================

st.markdown(
    '<div class="section-header">📧 IMPORTANT FINDINGS — EMAIL READY</div>',
    unsafe_allow_html=True
)

st.caption(
    "This section automatically generates an email summary based on the current "
    "Month / Week / Day / Issue Type filters. The main findings use only CG "
    "Helpline tickets; transferred/current-other-queue data is shown separately "
    "from the raw Group field."
)

email_text = make_email_summary(
    filtered,
    selected_month,
    selected_week,
    selected_day,
    selected_issue
)

email_html = make_email_html(
    filtered,
    filtered_other_groups,
    selected_month,
    selected_week,
    selected_day,
    selected_issue
)

st.info(
    "Copy the formatted email below with Ctrl+A / Ctrl+C and paste it into "
    "Outlook, Gmail or Excel. Tables remain tables and paste into separate "
    "Excel cells."
)

st.markdown(
    email_html,
    unsafe_allow_html=True
)


# ============================================================
# EXPORT
# ============================================================

st.markdown(
    '<div class="section-header">EXPORT</div>',
    unsafe_allow_html=True
)

csv_data = filtered.to_csv(
    index=False
).encode("utf-8")

st.download_button(
    "Download Filtered Ticket Data",
    data=csv_data,
    file_name="filtered_ticket_data.csv",
    mime="text/csv"
)


# ============================================================
# DOWNLOAD PROFESSIONAL EXCEL REPORT
# ============================================================

st.markdown(
    '<div class="section-header">📊 SHAREABLE EXCEL REPORT</div>',
    unsafe_allow_html=True
)

st.write(
    "Download a professionally formatted Excel report containing the "
    "original raw data, filterable analysis, KPI dashboard, L1/L2 issue "
    "breakup, top/low contributors, group analysis, and monthly/weekly/daily views."
)

excel_filter_text = filter_description(
    selected_month,
    selected_week,
    selected_day,
    selected_issue
)

# Legacy L1 summary used by the Excel dashboard sheet.
# The detailed L1 + L2 analysis is also exported separately.
issue_summary = summary_by_group(
    filtered,
    "Issue L1"
)

excel_bytes = build_excel_report(
    raw_data=df,
    filtered_data=filtered,
    issue_summary=issue_summary,
    month_summary=month_summary,
    week_summary=week_summary,
    day_summary=day_summary,
    email_text=email_text,
    filter_text=excel_filter_text,
)

st.download_button(
    label="⬇️ Download Color-Coded Excel Report",
    data=excel_bytes,
    file_name="CG_Helpline_Ticket_Analysis_Report.xlsx",
    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    type="primary",
)

st.caption(
    "The Excel report keeps the raw export on a separate Raw Data sheet "
    "and adds filterable analysis sheets without changing the original raw columns."
)


# ============================================================
# RAW DATA PREVIEW
# ============================================================

with st.expander("View Raw Data"):

    st.dataframe(
        df,
        use_container_width=True,
        height=400
    )
